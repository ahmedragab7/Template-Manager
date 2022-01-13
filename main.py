# This is a sample Python script.
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
# os.system("start EXCEL.EXE folder.xlsx")

import os
import tkinter
import openpyxl
import PIL.Image
from PIL import ImageTk
from tkinter import *
import win32com.client as win32
from openpyxl import load_workbook

user=0
passw=0
areanum=0
project_name=''
PDPNum=0
PPNum=0
list=[]
dev_names=[]
phase1_dev_g=[]
phase2_dev_g=[]
phase3_dev_g=[]

def devices_to_values(argument):
    switcher = {
        "EZC":1,
        "AUX":0.85,
        "HMI":1.5,
        "PLC":4.7,
        "SCANNER":0.8,
        "REPEATER":2,
        "EAG":4,
        "ENET_8":0.85,
        "ENET_24":0.8,
        "ENET_48":0.8
    }
    return switcher.get(argument, "nothing")


def main_L( list_local):
    global list
    global phase1_dev_g
    global phase2_dev_g
    global phase3_dev_g
    total_current=sum(list)
    num_of_dev=len(list)
    phase_current= total_current / 3
    phase1_curr=0
    phase1_dev=[]
    phase2_curr=0
    phase2_dev=[]
    phase3_curr=0
    phase3_dev=[]
    for dev in range(num_of_dev):

        taken=False
        d1=phase_current-phase1_curr+list[dev]
        d2=phase_current-phase2_curr+list[dev]
        d3=phase_current-phase3_curr+list[dev]
        if(d1>=d2 and d1>=d3 and taken==False):
            taken=True
            phase1_curr=phase1_curr+list[dev]
            phase1_dev.append(dev)
        if(d2>=d1 and d2>=d3 and taken==False):
            taken=True
            phase2_curr=phase2_curr+list[dev]
            phase2_dev.append(dev)
        if(d3>=d2 and d3>=d1 and taken==False):
            taken=True
            phase3_curr=phase3_curr+list[dev]
            phase3_dev.append(dev)

    print("ph1:",phase1_dev)
    print("ph2:",phase2_dev)
    print("ph3:",phase3_dev)
    phase1_dev_g=phase1_dev
    phase2_dev_g=phase2_dev
    phase3_dev_g=phase3_dev

def ok():
    global list
    global  dev_names
    print(devices_to_values(variable.get()))
    list.append(devices_to_values(variable.get()))
    dev_names.append(variable.get())
    list.sort(reverse=True)
    print("List_S:",list)
    main_L(list)
    text = Text(window1,bg="white",spacing2=1,height=20,width=10,x=100, y=415)
    text.place(x=290, y=315)
    for d in range(len(dev_names)):
        text.insert(INSERT,dev_names[d]+ '\n')

    #text.pack()

def area(en1):
    global areanum
    areanum=en1.widget.get()
    print(en1.widget.get())

def project(en1):
    global project_name
    project_name=en1.widget.get()
    print(en1.widget.get())

def PDPnum(en2):
    global PDPNum
    PDPNum=en2.widget.get()
    print(PDPNum)
def PPnum(en2):
    global PPNum
    PPNum=en2.widget.get()
    print(PPNum)

def PDP_Open():
    global areanum
    global PDPNum
    global project_name
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    proj_name=desktop+"/"+str(project_name)
    area_name="CA_"+str(areanum)
    PDP_name=area_name+"-"+"PDP_"+PDPNum

    if (os.path.exists(proj_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y =310)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name)
    if (os.path.exists(proj_name+"/"+area_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y =350)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name+"/"+area_name)

    name=str(PDP_name)+".xlsx"
    print("name:")
    print(name)
    if (os.path.exists(proj_name+"/"+area_name+"/"+name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y =390)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        wb = load_workbook('templates/PDP.xlsx')
        full_save_dir=proj_name+"/"+area_name+"/"+name
        wb.save(filename = full_save_dir)

    full_path=proj_name+"/"+area_name+"/"+name
    print(full_path)
    os.system("start EXCEL.EXE "+full_path)


def PP_Open():
    global areanum
    global PPNum
    global project_name
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    proj_name=desktop+"/"+str(project_name)
    area_name="CA_"+str(areanum)
    PP_name=area_name+"-"+"PP_"+PPNum
    if (os.path.exists(proj_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y =310)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name)
    if (os.path.exists(proj_name+"/"+area_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y = 350)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name+"/"+area_name)

    name=str(PP_name)+".xlsx"
    print("name:")
    print(name)
    if (os.path.exists(proj_name+"/"+area_name+"/"+name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y = 390)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        wb = load_workbook('templates/PP.xlsx')
        full_save_dir=proj_name+"/"+area_name+"/"+name
        wb.save(filename = full_save_dir)

    full_path=proj_name+"/"+area_name+"/"+name
    print(full_path)
    os.system("start EXCEL.EXE "+full_path)


    #for wb in excel.Workbooks:
    #    print("WB:",wb.Name)
    #    if (wb.Name == "bgarab.xlsx"):
    #        print("found")
    #        wb.SaveAs(name)


def DPS_Open():
    global areanum
    global project_name
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    proj_name=desktop+"/"+str(project_name)
    area_name="CA_"+str(areanum)
    DPS_name=area_name+"-"+"DPS_Calculation"
    if (os.path.exists(proj_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y =310)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name)
    if (os.path.exists(proj_name+"/"+area_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y = 350)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name+"/"+area_name)

    name=str(DPS_name)+".xlsx"
    print("name:")
    print(name)
    if (os.path.exists(proj_name+"/"+area_name+"/"+name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y = 390)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        wb = load_workbook('templates/DPS.xlsx')
        full_save_dir=proj_name+"/"+area_name+"/"+name
        wb.save(filename = full_save_dir)

    full_path=proj_name+"/"+area_name+"/"+name
    print(full_path)
    os.system("start EXCEL.EXE "+full_path)

def Alias_Open():
    global areanum
    global project_name
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    proj_name=desktop+"/"+str(project_name)

    area_name="CA_"+str(areanum)
    Alias_name=area_name+"-"+"Alias_Builder"
    if (os.path.exists(proj_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y =310)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name)
    if (os.path.exists(proj_name+"/"+area_name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y = 310)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        os.mkdir(proj_name+"/"+area_name)

    name=str(Alias_name)+".xlsx"
    print("name:")
    print(name)
    if (os.path.exists(proj_name+"/"+area_name+"/"+name)):
        tkinter.Label(window1, text = "Already exist", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 450,y = 350)#'username' is placed on position 00 (row - 0 and column - 0)
    else:
        wb = load_workbook('templates/Alias.xlsm')
        full_save_dir=proj_name+"/"+area_name+"/"+name
        wb.save(filename = full_save_dir)

    full_path=proj_name+"/"+area_name+"/"+name
    print(full_path)
    os.system("start EXCEL.EXE "+full_path)


def Result():
    global dev_names
    global list
    global phase1_dev_g
    global phase2_dev_g
    global phase3_dev_g
    CKT1 = Text(window1,bg="blue",fg="white",spacing2=1,height=10,width=20,x=300, y=400)
    CKT2 = Text(window1,bg="red",fg="white",spacing2=1,height=10,width=20,x=300, y=510)
    CKT3 = Text(window1,bg="yellow",spacing2=1,height=10,width=20,x=300, y=620)
    CKT1.place(x=390, y=300)
    CKT2.place(x=390, y=400)
    CKT3.place(x=390, y=500)
    c_sum=0
    for d in phase1_dev_g:
        CKT1.insert(INSERT,dev_names[d]+ '\n')
        c_sum+=list[d]
    CKT1.insert(INSERT,"Total current=")
    CKT1.insert(INSERT,c_sum)
    c_sum=0
    for d in phase2_dev_g:
        CKT2.insert(INSERT,dev_names[d]+ '\n')
        c_sum+=list[d]
    CKT2.insert(INSERT,"Total current=")
    CKT2.insert(INSERT,c_sum)
    c_sum=0
    for d in phase3_dev_g:
        CKT3.insert(INSERT,dev_names[d]+ '\n')
        c_sum+=list[d]
    CKT3.insert(INSERT,"Total current=")
    CKT3.insert(INSERT,c_sum)


def PP_dis():

    variable.set(devices[0]) # default value


    tkinter.OptionMenu(window1,variable, *[option for option in devices]).place(x=90,y=350)
    button = tkinter.Button(window1,text="ADD", command =ok,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',11),activeforeground ='black')
    button.place(x = 190, y = 350)
    button = tkinter.Button(window1,text="distribute", command =Result,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',11),activeforeground ='black')
    button.place(x = 190, y = 400)
    button_widget2 = tkinter.Button(window1,text="Close", command =Window2,activebackground = 'white',fg = 'white',bg ='red',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 190, y = 470)

def PDP():

    tkinter.Label(window1, text = "Number of the area :", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =350)#'username' is placed on position 00 (row - 0 and column - 0)
    tkinter.Label(window1, text = "PDP number :",fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =390) #'username' is placed on position 00 (row - 0 and column - 0)
    en1=tkinter.Entry(window1)
    en1.place(x =309, y = 350) # first input-field is placed on position 01 (row - 0 and column - 1)
    en1.bind("<Return>",area)
    tkinter.Label(window1, text = "project name :",fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =310) #'username' is placed on position 00 (row - 0 and column - 0)
    en3=tkinter.Entry(window1)
    en3.place(x =309, y = 310) # first input-field is placed on position 01 (row - 0 and column - 1)
    en3.bind("<Return>",project)
    en2=tkinter.Entry(window1)
    en2.place(x = 309, y = 390) # first input-field is placed on position 01 (row - 0 and column - 1)
    en2.bind("<Return>",PDPnum)
    button_widget2 = tkinter.Button(window1,text="Open", command =PDP_Open,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 10, y = 300)
    button_widget2 = tkinter.Button(window1,text="Close", command =Window2,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 60, y = 450)

def PP():

    tkinter.Label(window1, text = "Number of the area :", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =350)#'username' is placed on position 00 (row - 0 and column - 0)
    tkinter.Label(window1, text = "PP number :",fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =390) #'username' is placed on position 00 (row - 0 and column - 0)
    en1=tkinter.Entry(window1)
    en1.place(x =309, y = 350) # first input-field is placed on position 01 (row - 0 and column - 1)
    en1.bind("<Return>",area)
    tkinter.Label(window1, text = "project name :",fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =310) #'username' is placed on position 00 (row - 0 and column - 0)
    en3=tkinter.Entry(window1)
    en3.place(x =309, y = 310) # first input-field is placed on position 01 (row - 0 and column - 1)
    en3.bind("<Return>",project)
    en2=tkinter.Entry(window1)
    en2.place(x = 309, y = 390) # first input-field is placed on position 01 (row - 0 and column - 1)
    en2.bind("<Return>",PPnum)
    button_widget2 = tkinter.Button(window1,text="Open", command =PP_Open,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 10, y = 300)
    button_widget2 = tkinter.Button(window1,text="Close", command =Window2,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 60, y = 450)


def DPS():
    tkinter.Label(window1, text = "Number of the area :", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =350)#'username' is placed on position 00 (row - 0 and column - 0)
    en1=tkinter.Entry(window1)
    en1.place(x =309, y = 350) # first input-field is placed on position 01 (row - 0 and column - 1)
    en1.bind("<Return>",area)

    en2=tkinter.Entry(window1)
    tkinter.Label(window1, text = "project name :",fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =310) #'username' is placed on position 00 (row - 0 and column - 0)
    en3=tkinter.Entry(window1)
    en3.place(x =309, y = 310) # first input-field is placed on position 01 (row - 0 and column - 1)
    en3.bind("<Return>",project)
    button_widget2 = tkinter.Button(window1,text="Open", command =DPS_Open,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 10, y = 300)
    button_widget2 = tkinter.Button(window1,text="Close", command =Window2,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 60, y = 450)

def Alias():
    tkinter.Label(window1, text = "Number of the area :", fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =350)#'username' is placed on position 00 (row - 0 and column - 0)
    en1=tkinter.Entry(window1)
    en1.place(x =309, y = 350) # first input-field is placed on position 01 (row - 0 and column - 1)
    en1.bind("<Return>",area)

    en2=tkinter.Entry(window1)
    tkinter.Label(window1, text = "project name :",fg ="green" ,bg ='white',font =('BOLD',15)).place(x = 100,y =310) #'username' is placed on position 00 (row - 0 and column - 0)
    en3=tkinter.Entry(window1)
    en3.place(x =309, y = 310) # first input-field is placed on position 01 (row - 0 and column - 1)
    en3.bind("<Return>",project)
    button_widget2 = tkinter.Button(window1,text="Open", command =Alias_Open,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 10, y = 300)
    button_widget2 = tkinter.Button(window1,text="Close", command =Window2,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 60, y = 450)
def Window2():
    lable=tkinter.Label(window1,image=win2)
    lable.place(x=-40,y=-20)
    lable2=tkinter.Label(window1,image=ad)
    lable2.place(x=730,y=10)
    button_widget = tkinter.Button(window1,text="Open PDP FLA_Calculation", command = PDP,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget.place(x = 10, y = 55)
    button_widget = tkinter.Button(window1,text="Open PP FLA_Calculation", command = PP,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget.place(x = 10, y = 120)
    button_widget = tkinter.Button(window1,text="Open DPS Calculation", command = DPS,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget.place(x = 10, y = 185)
    button_widget = tkinter.Button(window1,text="Open Alias Builder", command = Alias,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget.place(x = 10, y = 250)
    button_widget = tkinter.Button(window1,text="Open Phase distributor", command = PP_dis,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget.place(x = 330, y = 55)
    button_widget2 = tkinter.Button(window1,text="Quit", command =window1.destroy,activebackground = 'white',fg = 'red',bg ='green',highlightcolor ='white',font =('BOLD',17),activeforeground ='black')
    button_widget2.place(x = 850, y = 600)

def login():
    global user,passw,var1
    if (user==1 and passw==1):
        print('pass')
        Window2()

def fun(en1):
    global user
    if (en1.widget.get()=="ahmed"):
        user=1
    print(en1.widget.get())

def fun2(en2):
    global passw
    if (en2.widget.get()=="000"):
        passw=1
    print(en2.widget.get())





def Start_program():

    var1 = IntVar()
    var2 = IntVar()

    lable=tkinter.Label(window1,image=im)
    lable.place(x=-40,y=-20)
    lable2=tkinter.Label(window1,image=ad)
    lable2.place(x=730,y=10)
    lable3=tkinter.Label(window1,image=am)
    lable3.place(x=900,y=630)
    # Use a breakpoint in the code line below to debug your script.
    tkinter.Label(window1, text = "Username", fg ="green" ,bg ='white',font =('BOLD',17)).place(x = 10,y =10)#'username' is placed on position 00 (row - 0 and column - 0)
    tkinter.Label(window1, text = "Password",fg ="green" ,bg ='white',font =('BOLD',17)).place(x = 10,y =50) #'username' is placed on position 00 (row - 0 and column - 0)
    en1=tkinter.Entry(window1)
    en1.place(x = 170, y = 10) # first input-field is placed on position 01 (row - 0 and column - 1)
    en1.bind("<Return>",fun)

    en2=tkinter.Entry(window1)
    en2.place(x = 170, y = 50) # first input-field is placed on position 01 (row - 0 and column - 1)
    en2.bind("<Return>",fun2)
    button_widget = tkinter.Button(window1,text="Login", command = login,activebackground = 'white',fg = 'white',bg ='green',highlightcolor ='white',font ='BOLD',activeforeground ='black')
    button_widget.place(x = 10, y = 110)
    window1.mainloop()


def bgrab():
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    for wb in excel.Workbooks:
        print("WB:",wb.Name)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    window1=tkinter.Tk()
    window1.geometry("1000x700")
    window1.title("Templates manager ")
    i=PIL.Image.open("main_window.jpg")
    im=ImageTk.PhotoImage(i)
    i2=PIL.Image.open("Advansys.png")
    ad=ImageTk.PhotoImage(i2)
    i3=PIL.Image.open("am.png")
    am=ImageTk.PhotoImage(i3)
    i4=PIL.Image.open("window2.jpg")
    win2=ImageTk.PhotoImage(i4)
    devices=["EZC", "AUX", "HMI","PLC","SCANNER","REPEATER","EAG","ENET_8","ENET_24","ENET_48"]
    variable = tkinter.StringVar(window1)
    #excel = win32.gencache.EnsureDispatch('Excel.Application')
    Start_program()
    bgrab()



# See PyCharm help at https://www.jetbrains.com/help/pycharm/
